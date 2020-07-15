# -*- coding: utf-8 -*-

import io
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

import json
import __init__
import utils
import const as CONST

import datetime

import win32com.client
from win32com.client import Dispatch

def get_xlsx_raw(xlsx, sheet, begin=1, end=sys.maxsize, headers={}):
    '''Get raw data of table from excel.'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get raw data from %s %s", Path(xlsx).name, sheet)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]

        data = [[val(cell) for cell in row] for row in wb[sheet].rows]
        data = data[begin-1:min(end, sys.maxsize)]

        first_row = data[0][:]
        data[0] = [headers.get(col, col) for col in data[0]]

        if headers != {}:
            data.append(first_row)

    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data


def get_xlsx_cells(xlsx, sheet, list_cell):
    '''Get cell value from excel file'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get value of cell %s", list_cell)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]
        ws = wb[sheet]

        data = {key: val(ws[key]) for key in list_cell}
    except Exception as e:
        logger.exception(e)
        data = {}
    finally:
        wb.close()
        return data


def get_xlsx_sheets(xlsx):
    '''Get sheets of xlsx'''
    logger.debug("Get sheets from file %s", xlsx)
    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        data = wb.sheetnames
    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data

def parse_summary_json(file, sheetname="", begin=47, end=47):
    try:
        result = dict()

        if sheetname == "":
            sheetname = utils.load(CONST.SETTING).get("sheetname")
    
        sheet_name = [str(x) for x in get_xlsx_sheets(file) if x == sheetname]

        header = [h for h in get_xlsx_raw(file, sheet_name[0], begin=begin, end=begin)[0]]
        data = get_xlsx_raw(file, sheet_name[0], begin=begin+1, end=end)
        d_data = dict()
        result = dict()
        for index, d in enumerate(data):
            d_data = dict(zip(header, d))
            for key, val in dict(d_data).items():
                if key is None:
                    del d_data[key]
            result[index + 1] = dict(d_data)

    except Exception as e:
        logger.exception(e)
        result = {}
    finally:
        return dict(result)
